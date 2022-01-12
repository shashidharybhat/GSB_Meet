from os import close
import pandas as pd
import re
import openpyxl 
from difflib import get_close_matches
from openpyxl.styles import PatternFill, Font

opportunities = ['mitr', 'mitr', 'mirt', 'mitra', 'mithra', 'mithr', 'saathi', 'saathi', 'saatih', 'saahti', 'satahi', 'saathi', 'asathi', 'developer club', 'developers  club', 'developer club', 'developerclub', 'developer clbu', 'developer culb', 'developer lcub', 'developerc lub', 'develope rclub', 'developre club', 'develoepr club', 'develpoer club', 'deveolper club', 'devleoper club', 'deevloper club', 'dveeloper club', 'edveloper club', 'dev club', 'dev clbu', 'dev culb', 'dev lcub', 'devc lub', 'de vclub', 'dve club', 'edv club', 'editorial', 'editoria', 'editoril', 'edtorial', 'editoial', 'editrial', 'ediorial', 'edtorial', 'eitorial', 'ditorial', 'editorila', 'editorail', 'editoiral', 'editroial', 'ediotrial', 'edtiorial', 'eidtorial', 'deitorial', 'blog', 'blogg', 'bloog', 'bllog', 'bblog', 'blgo', 'bolg', 'lbog', 'gdsc', 'gdcs', 'gsdc', 'dgsc', 'regional', 'regiona', 'regionl', 'regioal', 'reginal', 'regonal', 'reional', 'rgional', 'egional', 'regionall', 'regionaal', 'regionnal', 'regioonal', 'regiional', 'reggional', 'reegional', 'rregional', 'regionla', 'regioanl', 'reginoal', 'regoinal', 'reigonal', 'rgeional', 'ergional', 'trv', 'trivikra', 'trivikrm', 'tivikram', 'triviram', 'trvikram', 'triikram', 'trvikram', 'tivikram', 'rivikram', 'trivikrma', 'trivikarm', 'trivirkam', 'trivkiram', 'triivkram', 'trviikram', 'tirvikram', 'rtivikram', 'tvr', 'rtv', 'ashwin', 'ashwi', 'ashwn', 'ashin', 'aswin', 'ahwin', 'shwin', 'ashwni', 'ashiwn', 'aswhin', 'ahswin', 'sahwin', 'ash', 'vue', 'veu', 'view', 'vvu', 'placement', 'placement', 'placemen', 'placemet', 'placment', 'placeent', 'placment', 'plaement', 'plcement', 'pacement', 'lacement', 'placemetn', 'placemnet', 'placeemnt', 'placmeent', 'plaecment', 'plcaement', 'palcement', 'lpacement', 'placements', 'ds', 'ds', 'bs', 'entrepreneurship', 'entrepeneurship', 'entrepreneur', 'entepreneur', 'entreprener', 'ntrepreneur', 'etrepreneur', 'ntrepreneur', 'entepreneur', 'entrereneur', 'ntrepreneur', 'entepreneur', 'enrepreneur', 'etrepreneur', 'ntrepreneur', 'entrepreneru', 'entreprenuer', 'entrepreenur', 'entreprneeur', 'entreperneur', 'entrerpeneur', 'entrpereneur', 'enterpreneur', 'enrtepreneur', 'etnrepreneur', 'netrepreneur', 'ecell', 'e-cell']
font = Font(name='Calibri',size=11,bold=False,italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                color='FF000000')
fill = PatternFill(fill_type='solid',
                 start_color='FFFFFFFF',
                 end_color='FF0000FF')
def mistypes(types):
    keywords = {"mitr":["mitr", "mirt", "mitra", "mithra", "mithr"], "saathi":["saathi","saatih", "saahti", "satahi", "saathi", "asathi"], "developer club":
["developers  club", "developer club", "developerclub", "developer clbu", "developer culb", "developer lcub", "developerc lub", "develope rclub", "developre club", "develoepr club", "develpoer club", "deveolper club", "devleoper club", "deevloper club", "dveeloper club", "edveloper club","dev club", "dev clbu", "dev culb", "dev lcub", "devc lub", "de vclub", "dve club", "edv club" ], "editorial":["editoria", "editoril", "edtorial", "editoial", "editrial", "ediorial", "edtorial", "eitorial", "ditorial", "editorila", "editorail", "editoiral", "editroial", "ediotrial", "edtiorial", "eidtorial", "deitorial"], "blog":["blogg", "bloog", "bllog", "bblog", "blgo", "bolg", "lbog"], "gdsc":["gdcs", "gsdc", "dgsc"], "regional":["regiona", "regionl", "regioal", "reginal", "regonal", "reional", "rgional", "egional", "regionall", "regionaal", "regionnal", "regioonal", "regiional", "reggional", "reegional", "rregional", "regionla", "regioanl", "reginoal", "regoinal", "reigonal", "rgeional", "ergional"], "trv":["trivikra", "trivikrm", "tivikram", "triviram", "trvikram", "triikram", "trvikram", "tivikram", "rivikram", "trivikrma", "trivikarm", "trivirkam", "trivkiram", "triivkram", "trviikram", "tirvikram", "rtivikram", "tvr", "rtv"], "ashwin":["ashwi", "ashwn", "ashin", "aswin", "ahwin", "shwin", "ashwni", "ashiwn", "aswhin", "ahswin", "sahwin", "ash"], "vue": ["veu", "view", "vvu"], "placement":["placement", "placemen", "placemet", "placment", "placeent", "placment", "plaement", "plcement", "pacement", "lacement", "placemetn", "placemnet", "placeemnt", "placmeent", "plaecment", "plcaement", "palcement", "lpacement", "placements"], "ds":["ds", "bs"], "entrepreneurship": ["entrepeneurship","entrepreneur","entepreneur", "entreprener", "ntrepreneur", "etrepreneur", "ntrepreneur", "entepreneur", "entrereneur", "ntrepreneur", "entepreneur", "enrepreneur", "etrepreneur", "ntrepreneur", "entrepreneru", "entreprenuer", "entrepreenur", "entreprneeur", "entreperneur", "entrerpeneur", "entrpereneur", "enterpreneur", "enrtepreneur", "etnrepreneur", "netrepreneur", "ecell","e-cell"]
 }
    for i,k in keywords.items():
        if types in k:
            return i.title()
    return None


def closeMatches(word):
    word = word.split(" ")
    matches = []
    for words in word:
        key = get_close_matches(words, opportunities, 3, 0.5)
        if key != []:
            matches.append(mistypes(key[0]))
    return matches
     
rx_dict = {
    'name': re.compile(r'from (?P<name>.*)to/i'),
    'keywords': re.compile(r'(?P<keywords>.*@student.onlinedegree.iitm.ac.in)'),
    'email': re.compile(r'(?P<email>2.*@student.onlinedegree.iitm.ac.in)'),
}
def _parse_line(line):
    """
    Do a regex search against all defined regexes and
    return the key and match result of the first matching regex

    """
    matches = []
    for key, rx in rx_dict.items():
        match = rx.search(line, re.IGNORECASE)
        if match:
            #print(key, " Hi ", match)
            return key, match
    # if there are no matches
    return None,None
    
def check_email(email):

# pass the regular expression
# and the string into the fullmatch() method
    if(re.fullmatch(r'\b2[A-Za-z0-9._%+-]+@student.onlinedegree.iitm.ac.in\b', email)):
        return email

    else:
        return ''

def parse_file(filepath):
    """
    Parse text at given filepath

    Parameters
    ----------
    filepath : str
        Filepath for file_object to be parsed

    Returns
    -------
    data : pd.DataFrame
        Parsed data

    """

    data = []  # create an empty list to collect the data
    # open the file and read through it line by line
    with open(filepath, 'r') as file_object:
        line = file_object.readline()
        while line:
            # at each line check for a match with a regex
            key, match = _parse_line(line)
            # extract name name
            if key == 'name':
                name = match.group('name')
                name = name.lower()
            # extract Email and Interests
            if key == 'keywords':
                dataLine = match.group('keywords')
                dataLine = dataLine.lower()
                email = dataLine[-42:]
                interests = closeMatches(dataLine[:-42])
                if interests != []:
                    row = {
                            'Email': email,
                        }
                    for i,inde in enumerate(["Interest 1", "Interest 2", "Interest 3", "Interest 4", "Interest 5", "Interest 6", "Interest 7", "Interest 8",  "Interest 9"]):
                        row[inde] = interests[i] if i < len(interests)-1 else None
                    # append the dictionary to the data list
                    data.append(row)
            if key == 'email':
                email = check_email(str(match.group('email')))
                #print(email)
            line = file_object.readline()

        # create a pandas DataFrame from the list of dicts
        data = pd.DataFrame(data)
        data.set_index(['Email'], inplace=True)
        # consolidate df to remove nans
        data = data.groupby(level=data.index.names).first()
        data = data.apply(pd.to_numeric, errors='ignore')
    return data
if __name__ == '__main__':
    filepath = 'GSB_Meet.txt'
    data = parse_file(filepath)
    data.to_excel("GSB_Data.xlsx")
    path = "GSB_Data.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj.active 
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 15
    sheet.column_dimensions['I'].width = 15
    sheet.column_dimensions['J'].width = 15
    wb_obj.save('GSB_Excel.xlsx') 
    
